import openpyxl as xl

# Document_path_one must be my document and the path two must be the data_base from where
# I´ll get the info


def modify_excel_doc(document_path_one, sheet_to_work_one, cell_to_evaluate_one,
                     document_path_two, sheet_to_work_two, cell_to_evaluate_two):

    # Opening booth documents
    my_document_one = xl.load_workbook(document_path_one)
    my_document_two = xl.load_workbook(document_path_two)

    # Here I see which sheet of each document I´ll be working on
    sheet_doc_one = my_document_one[sheet_to_work_one]
    sheet_doc_two = my_document_two[sheet_to_work_two]

    rows_number = sheet_doc_one.max_row

    for row in range(2, rows_number + 1):
        my_cell_doc_one = sheet_doc_one.cell(row, cell_to_evaluate_one)     # Coordinates of the cell

        # Here somehow I need to check if the name of the movie from my list is in the data_base
        if str(my_cell_doc_one.value).lower() == "keep later going":
            print(my_cell_doc_one.value)

        # What I have in mind is to create a dictionary with the name of the movies as a key
        # and the ages as their values, so if the value of the cell is in the dictionary,
        # I can get its value and add it to my document

